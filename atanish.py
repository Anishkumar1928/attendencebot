import openpyxl
from datetime import datetime
import itertools
def extrat_stu(file_path):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet1']
    totalstudent=[]    
    for row in range(2, sheet.max_row):
        if sheet.cell(row=row, column=1).value==None:
            continue
        else:
            totalstudent.append(sheet.cell(row=row, column=1).value)
    return totalstudent


def take_attendance(file_path,pr,ab):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['Sheet1']
    

    # Get today's date
    today = datetime.now().date()
    today_column =None

    for colum in itertools.count(start=2):
        if sheet.cell(row=1, column=colum).value==None:
             sheet.cell(row=1,column=colum,value=str(today))
             today_column=colum
             break
    for c in range(2, sheet.max_column +1):
        print(sheet.cell(row=1, column=c).value)    


    # Find the column corresponding to today's date
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value in ab:
            sheet.cell(row=row, column=today_column, value='absent')
        elif sheet.cell(row=row, column=1).value in pr:
            sheet.cell(row=row, column=today_column, value='present') 
    wb.save(file_path)
    print("Attendance saved.")



"""if __name__ == "__main__":
    pr=[]
    ab=["KRITI KUMARI","HIMANSHU KUMAR","HARSH RAJ","AAKASH RAJ","SAUMYA KUMARI","ANSHU KUMARI","PANKAJ KUMAR SAH"]"""
file_path = "attendance.xlsx"
    #take_attendance(file_path,pr,ab)"""
print(extrat_stu(file_path))